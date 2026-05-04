#!/usr/bin/env python3
"""Generate the Kinderhook Farmers' Market content management spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, timedelta

wb = openpyxl.Workbook()

# Shared styles
header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
orange_fill = PatternFill(start_color="F26A2A", end_color="F26A2A", fill_type="solid")
teal_fill = PatternFill(start_color="2AAFCE", end_color="2AAFCE", fill_type="solid")
cream_fill = PatternFill(start_color="FFF9F0", end_color="FFF9F0", fill_type="solid")
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
center_alignment = Alignment(horizontal="center", vertical="center")


def style_headers(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def style_data_rows(ws, num_cols):
    for row in ws.iter_rows(min_row=2, max_col=num_cols):
        for cell in row:
            cell.border = thin_border
            cell.alignment = wrap_alignment
            # Alternate row shading
            if cell.row % 2 == 0:
                cell.fill = cream_fill


# ─── Tab 1: Weekly Schedule ───

ws1 = wb.active
ws1.title = "Weekly Schedule"
ws1.sheet_properties.tabColor = "F26A2A"

headers_1 = [
    "Date",
    "Display Date",
    "Theme",
    "Note",
    "Music Act",
    "Music Time",
    "Special Event Name",
    "Special Event Time",
    "Special Event Description",
    "Guest Vendors",
    "Community Partners",
    "Regular Vendors",
    "Status",
]
ws1.append(headers_1)

# Column widths
widths_1 = [14, 28, 20, 18, 20, 20, 24, 16, 34, 34, 30, 34, 12]
for i, w in enumerate(widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# 2026 season music lineup (10am–12pm each Saturday).
music_lineup = {
    "2026-05-02": "Alley Cats",
    "2026-05-09": "Scott Stockman",
    "2026-05-16": "Mike Pagnani",
    "2026-05-23": "Alejandra Maciel & Friend",
    "2026-05-30": "Jasperoo",
    "2026-06-06": "The Sound of Somewhere",
    "2026-06-13": "Scott Stockman",
    "2026-06-20": "Alejandra Maciel & Friend",
    "2026-06-27": "Jasperoo",
    "2026-07-04": "Brad & Friends (official title to come)",
    "2026-07-11": "Scott Stockman",
    "2026-07-18": "The Last Pangeans",
    "2026-07-25": "Jasperoo",
    "2026-08-01": "Alley Cats",
    "2026-08-08": "Scott Stockman",
    "2026-08-15": "Too Lazy Boys",
    "2026-08-22": "Jasperoo",
    "2026-08-29": "The Sound of Somewhere",
    "2026-09-05": "Mike Pagnani",
    "2026-09-12": "Scott Stockman",
    "2026-09-19": "Jasperoo",
    "2026-09-26": "Alley Cats",
    "2026-10-03": "Mike Pagnani",
    "2026-10-10": "Scott Stockman & Friends",
    "2026-10-17": "Monty Bopp",
    "2026-10-24": "Alejandra Maciel & Friend",
    "2026-10-31": "TC",
}

# Pre-fill Saturdays for 2026 season (May 2 - Oct 31)
season_start = date(2026, 5, 2)
season_end = date(2026, 10, 31)
current = season_start
row_num = 2
while current <= season_end:
    iso = current.strftime("%Y-%m-%d")
    ws1.cell(row=row_num, column=1, value=iso)
    # Display date
    ws1.cell(
        row=row_num,
        column=2,
        value=current.strftime("%A, %B %-d, %Y"),
    )
    # Default note
    ws1.cell(row=row_num, column=4, value="Rain or Shine")
    # Pre-fill music from 2026 lineup
    if iso in music_lineup:
        ws1.cell(row=row_num, column=5, value=music_lineup[iso])
        ws1.cell(row=row_num, column=6, value="10:00am - 12:00pm")
    # Default status
    ws1.cell(row=row_num, column=13, value="Draft")
    current += timedelta(weeks=1)
    row_num += 1

# 2026 weekly vendors (present every Saturday). Source: weekly-vendors.json.
weekly_vendors_list = (
    "Cedar Ridge Farm Ghent, Damsel Garden, Glencadia Greenhouses, "
    "Heritage Hill Pork (Formerly Lovers Leap Farm), Hickory Creek Farms, "
    "Mort's Maple, River House Market, "
    "Albany Distilling, Hamrah's Lebanese Foods & Takeaway, "
    "Our Old House Bakery, Worldling's Pleasure, "
    "Kipling & Raven Dog Treats, O'Boys Soaps, "
    "Valatie Food Pantry"
)
for row in range(2, row_num):
    ws1.cell(row=row, column=12, value=weekly_vendors_list)

# Pre-fill per-date themes, guest vendors, special events, and extended hours.
# Special-event headlines pulled from src/content/events/. The full event
# inventory lives on the Village Events tab; one headline per market day here.
weekly_overrides = {
    "2026-05-02": {
        "theme": "Opening Day — Early Spring Market",
        "guest_vendors": "Columbia Friends of the Electric Trail, Wild Lavender Crafts",
        "special_name": "Season Opening Celebration",
        "special_time": "8:30am",
        "special_desc": "Welcome back the market for the 2026 season! (Library: Stories for Pups, 1–2pm.)",
        "status": "Ready",
    },
    "2026-05-09": {
        "special_name": "Super Stories Open Maker Hours — Mother's Day Card Making",
        "special_time": "10:00am - 12:00pm",
        "special_desc": "Drop in to Super Stories and make a card for Mother's Day. (Library: Travel Planning Made Easy, 2–3pm.)",
    },
    "2026-05-16": {
        "special_name": "Festival of the Unknown",
        "special_time": "11:00am - 2:00pm",
        "special_desc": "Library lectures on cryptozoology, ufology, parapsychology, plus authors and a documentary on local cryptids.",
    },
    "2026-05-23": {
        "special_name": "Village-Wide Spring Yard Sale",
        "special_time": "9:00am until sold out",
        "special_desc": "Treasures scattered across Kinderhook.",
    },
    "2026-05-30": {
        "theme": "Kinderhook Makers Market — Extended Market Day",
        "note": "Extended hours: 8:30am - 2:00pm. Rain or Shine",
        "special_name": "Kinderhook Makers Market / Fyfe & Drumms Muster & Parade / Modus Operandi Opening",
        "special_time": "8:30am - 2:00pm (market); Parade noon; Gallery 2pm",
        "special_desc": "Extended market day. Fyfe & Drumms Muster & Parade on Broad Street at noon. Modus Operandi opens at The School (Jack Shainman) at 2pm.",
    },
    "2026-06-06": {
        "special_name": "Seen Scenes Opening / OK 5K / Persons of Color Cemetery Tour",
        "special_time": "Reception 3–5pm; 5K & tour times TBD",
        "special_desc": "Seen Scenes opening reception at Kinderhook Knitting Mill. OK 5K in the village. Persons of Color Cemetery Tour. (Library: Stories for Pups, 1–2pm.)",
    },
    "2026-06-13": {
        "special_name": "Volunteer Fair",
        "special_time": "10:00am - 2:00pm",
        "special_desc": "Library's inaugural volunteer fair — meet local nonprofits.",
    },
    "2026-06-20": {
        "special_name": "Rising Star Dance Academy Performance",
        "special_time": "11:00am",
        "special_desc": "Live performance at the market.",
    },
    "2026-06-27": {
        "special_name": "Kinderhook Pride Parade",
        "special_time": "2:00pm",
        "special_desc": "Hudson Street to Kinderhook Village Square.",
    },
    "2026-07-04": {
        "theme": "July 4th — Extended Market Day",
        "note": "Extended hours: 8:30am - 1:30pm. Rain or Shine",
        "special_name": "KBPA People's Parade",
        "special_time": "11:30am",
        "special_desc": "Parade kicks off from Rothermel Park. Market runs extended hours to 1:30pm.",
    },
    "2026-09-19": {
        "special_name": "Village-Wide Fall Yard Sale",
        "special_time": "9:00am until sold out",
        "special_desc": "Treasures scattered across Kinderhook.",
    },
    "2026-10-10": {
        "theme": "Fall Festival — Extended Market Day",
        "note": "Extended hours: 8:30am - 2:00pm. Rain or Shine",
        "special_name": "Fall Festival & Kinderhook Makers Market",
        "special_time": "8:30am - 2:00pm",
        "special_desc": "Extended market day with the Kinderhook Makers Market.",
    },
    "2026-10-31": {
        "theme": "Final Market Day of 2026 Season",
    },
}

# Apply overrides.
for row in range(2, row_num):
    iso = ws1.cell(row=row, column=1).value
    override = weekly_overrides.get(iso)
    if not override:
        continue
    if "theme" in override:
        ws1.cell(row=row, column=3, value=override["theme"])
    if "note" in override:
        ws1.cell(row=row, column=4, value=override["note"])
    if "special_name" in override:
        ws1.cell(row=row, column=7, value=override["special_name"])
    if "special_time" in override:
        ws1.cell(row=row, column=8, value=override["special_time"])
    if "special_desc" in override:
        ws1.cell(row=row, column=9, value=override["special_desc"])
    if "guest_vendors" in override:
        ws1.cell(row=row, column=10, value=override["guest_vendors"])
    if "status" in override:
        ws1.cell(row=row, column=13, value=override["status"])

# Status dropdown validation
status_val = DataValidation(type="list", formula1='"Draft,Ready,Published"', allow_blank=True)
status_val.error = "Please select Draft, Ready, or Published"
status_val.errorTitle = "Invalid Status"
ws1.add_data_validation(status_val)
for row in range(2, row_num):
    status_val.add(ws1.cell(row=row, column=13))

style_headers(ws1, len(headers_1))
style_data_rows(ws1, len(headers_1))

# Conditional color for status column
for row in range(2, row_num):
    cell = ws1.cell(row=row, column=13)
    cell.alignment = center_alignment
    val = cell.value
    if val == "Ready":
        cell.fill = yellow_fill
    elif val == "Published":
        cell.fill = green_fill


# ─── Tab 2: Vendors ───

ws2 = wb.create_sheet("Vendors")
ws2.sheet_properties.tabColor = "B8BF3D"

headers_2 = [
    "Name",
    "Type",
    "Tagline",
    "Categories",
    "Description",
    "Website",
    "Instagram",
    "Facebook",
    "Featured",
    "Active",
]
ws2.append(headers_2)

widths_2 = [24, 14, 28, 24, 40, 30, 20, 30, 12, 10]
for i, w in enumerate(widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Vendors. Source: src/content/vendors/*.md (entries with profile pages) +
# src/data/weekly-vendors.json (weekly regulars without profile pages).
# Type maps vendorType=regular → "Weekly", vendorType=guest → "Rotating".
vendors = [
    # ── Weekly regulars without profile pages (from weekly-vendors.json) ──
    ["Albany Distilling", "Weekly", "", "spirits, food & drink", "", "", "", "", "No", "Yes"],
    ["Cedar Ridge Farm Ghent", "Weekly", "", "produce, farm", "", "", "", "", "No", "Yes"],
    ["Glencadia Greenhouses", "Weekly", "", "plants, produce, farm", "", "", "", "", "No", "Yes"],
    ["Hamrah's Lebanese Foods & Takeaway", "Weekly", "", "prepared foods, food & drink", "", "", "", "", "No", "Yes"],
    ["Heritage Hill Pork (Formerly Lovers Leap Farm)", "Weekly", "", "meat, farm", "", "", "", "", "No", "Yes"],
    ["Hickory Creek Farms", "Weekly", "", "produce, farm", "", "", "", "", "No", "Yes"],
    ["Kipling & Raven Dog Treats", "Weekly", "", "pet, home & body", "", "", "", "", "No", "Yes"],
    ["Mort's Maple", "Weekly", "", "maple, farm", "", "", "", "", "No", "Yes"],
    ["O'Boys Soaps", "Weekly", "", "soap, home & body", "", "", "", "", "No", "Yes"],
    ["Our Old House Bakery", "Weekly", "", "baked goods, food & drink", "", "", "", "", "No", "Yes"],
    ["Valatie Food Pantry", "Weekly", "", "community", "", "", "", "", "No", "Yes"],
    ["Worldling's Pleasure", "Weekly", "", "food & drink", "", "", "", "", "No", "Yes"],

    # ── Vendors with profile pages (from src/content/vendors/) ──
    ["Autumn's Essentials", "Rotating", "Natural wellness products",
     "wellness, body-care, handmade",
     "Handcrafted soaps, lotions, and wellness products made with natural ingredients.",
     "", "", "", "No", "No"],
    ["Courtney Aison Pottery", "Rotating", "Functional ceramics for everyday life",
     "crafts, pottery, home-goods",
     "Handmade functional pottery: mugs, bowls, plates, and vases. Wheel-thrown and uniquely glazed.",
     "", "", "", "No", "No"],
    ["Crème de la Crème Bakery", "Rotating", "Mother/daughter bakers and entrepreneurs",
     "baked-goods, sweets",
     "A mother-daughter baking duo bringing artisan pastries, fresh bread, and sweet treats.",
     "", "", "", "Yes", "No"],
    ["Damsel Garden", "Weekly", "Sweet peas, peonies & seasonal blooms",
     "flowers, plants, bouquets",
     "Locally grown cut flowers and seasonal bouquets from Columbia County.",
     "", "", "", "Yes", "No"],
    ["Grandaddy Weaves Honey", "Weekly", "Local honey from happy bees",
     "honey, specialty-foods",
     "Pure, raw honey harvested from hives in the Hudson Valley.",
     "", "", "", "Yes", "No"],
    ["Oona Montalvo Pottery", "Rotating", "Handcrafted ceramics for everyday beauty",
     "crafts, pottery, home-goods",
     "Handmade pottery: mugs, bowls, vases, and decorative pieces. Fired locally.",
     "", "", "", "No", "No"],
    ["Peta's Pocket Caribbean BBQ", "Rotating", "Caribbean flavors in the Hudson Valley",
     "prepared-foods, bbq, international",
     "Authentic Caribbean BBQ. Jerk chicken, rice and peas, plantains, and more.",
     "", "", "", "Yes", "No"],
    ["Reagan Rose Chocolate-Chip Cookies", "Rotating", "Homemade cookies with love",
     "baked-goods, sweets, cookies",
     "Classic chocolate chip cookies and other homemade treats baked fresh for market day.",
     "", "", "", "No", "No"],
    ["River House Market", "Weekly", "Farm fresh from the Hudson",
     "produce, farm",
     "Fresh seasonal produce grown along the Hudson River. Vegetables, herbs, and specialty crops.",
     "", "", "", "Yes", "No"],
    ["Samascott Orchards", "Weekly", "Six generations of family farming",
     "produce, fruit, cider",
     "Hudson Valley institution since 1821. Seasonal fruits, vegetables, cider, and farm-fresh eggs.",
     "https://samascottorchards.com", "@samascottorchards", "", "Yes", "No"],
    ["Wyrm Farm", "Weekly", "Regenerative farming for a better future",
     "produce, farm, vegetables",
     "Vegetables grown using regenerative farming practices. Building healthy soil for healthy food.",
     "", "", "", "Yes", "No"],
]
for v in vendors:
    ws2.append(v)

# Type dropdown
type_val = DataValidation(type="list", formula1='"Weekly,Rotating"', allow_blank=True)
ws2.add_data_validation(type_val)
for row in range(2, 2 + len(vendors) + 20):  # extra rows for future entries
    type_val.add(ws2.cell(row=row, column=2))

# Featured/Active dropdowns
yn_val = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws2.add_data_validation(yn_val)
for row in range(2, 2 + len(vendors) + 20):
    yn_val.add(ws2.cell(row=row, column=9))
    yn_val.add(ws2.cell(row=row, column=10))

style_headers(ws2, len(headers_2))
style_data_rows(ws2, len(headers_2))


# ─── Tab 3: Recipes ───

ws3 = wb.create_sheet("Recipes")
ws3.sheet_properties.tabColor = "2AAFCE"

headers_3 = [
    "Title",
    "Contributor",
    "Contributor Link",
    "Prep Time",
    "Cook Time",
    "Servings",
    "Difficulty",
    "Season",
    "Produce Used",
    "Tags",
    "Ingredients",
    "Instructions",
    "Status",
]
ws3.append(headers_3)

widths_3 = [28, 20, 26, 14, 14, 10, 12, 12, 24, 24, 40, 50, 12]
for i, w in enumerate(widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Recipes. Source: src/content/recipes/*.md. Ingredients/Instructions
# bodies live in the markdown files; metadata only is mirrored here so the
# coordinator can sweep over status without having to open each file.
recipes = [
    ["Spring Asparagus Salad with Lemon Vinaigrette", "Market Kitchen", "",
     "15 minutes", "5 minutes", 4, "easy", "spring",
     "asparagus, radishes, spring greens",
     "vegetarian, gluten-free, quick, spring",
     "See src/content/recipes/spring-asparagus-salad.md",
     "See src/content/recipes/spring-asparagus-salad.md",
     "Published"],
    ["Samascott Orchard Apple Crisp", "The Samascott Family", "",
     "20 min", "45 min", 8, "easy", "fall",
     "Apples, Oats",
     "dessert, fall-favorite, comfort-food, vegetarian",
     "See src/content/recipes/samascott-apple-crisp.md",
     "See src/content/recipes/samascott-apple-crisp.md",
     "Published"],
    ["Peta's Market Day Jerk Chicken Bowl", "Peta, Peta's Pocket Caribbean BBQ", "",
     "30 min (plus marinating)", "25 min", 4, "medium", "summer",
     "Peppers, Tomatoes, Greens, Fresh Herbs",
     "main-dish, caribbean, grilled, gluten-free",
     "See src/content/recipes/petas-jerk-chicken-bowl.md",
     "See src/content/recipes/petas-jerk-chicken-bowl.md",
     "Published"],
    ["Kinderhook Dutch Baby Pancake", "Margaret Van Dyke, Kinderhook Resident", "",
     "10 min", "20 min", 4, "easy", "spring",
     "Eggs, Berries, Maple Syrup",
     "breakfast, dutch-heritage, weekend-brunch, vegetarian",
     "See src/content/recipes/kinderhook-dutch-pancake.md",
     "See src/content/recipes/kinderhook-dutch-pancake.md",
     "Published"],
    ["Grandaddy's Honey Glazed Carrots", "Dave from Grandaddy Weaves Honey", "",
     "10 min", "25 min", 4, "easy", "spring",
     "Carrots, Honey, Fresh Thyme",
     "side-dish, vegetarian, local-honey, kid-friendly",
     "See src/content/recipes/grandaddys-honey-glazed-carrots.md",
     "See src/content/recipes/grandaddys-honey-glazed-carrots.md",
     "Published"],
]
for r in recipes:
    ws3.append(r)

# Difficulty dropdown
diff_val = DataValidation(type="list", formula1='"easy,medium,hard"', allow_blank=True)
ws3.add_data_validation(diff_val)
for row in range(2, 50):
    diff_val.add(ws3.cell(row=row, column=7))

# Season dropdown
season_val = DataValidation(type="list", formula1='"spring,summer,fall,all"', allow_blank=True)
ws3.add_data_validation(season_val)
for row in range(2, 50):
    season_val.add(ws3.cell(row=row, column=8))

# Recipe status dropdown
recipe_status_val = DataValidation(type="list", formula1='"Draft,Ready,Published"', allow_blank=True)
ws3.add_data_validation(recipe_status_val)
for row in range(2, 50):
    recipe_status_val.add(ws3.cell(row=row, column=13))

style_headers(ws3, len(headers_3))
style_data_rows(ws3, len(headers_3))


# ─── Tab 4: Music Schedule ───

ws_music = wb.create_sheet("Music Schedule")
ws_music.sheet_properties.tabColor = "5D3C54"

headers_music = ["Date", "Day", "Performer", "Time", "Notes"]
ws_music.append(headers_music)
for i, w in enumerate([14, 20, 32, 22, 30], 1):
    ws_music.column_dimensions[get_column_letter(i)].width = w

for iso, act in music_lineup.items():
    d = date.fromisoformat(iso)
    note = "Official title to come" if "(official title" in act else ""
    act_clean = act.replace(" (official title to come)", "")
    ws_music.append([iso, d.strftime("%A, %B %-d"), act_clean, "10:00am - 12:00pm", note])

style_headers(ws_music, len(headers_music))
style_data_rows(ws_music, len(headers_music))


# ─── Tab 5: Village Events ───

ws_events = wb.create_sheet("Village Events")
ws_events.sheet_properties.tabColor = "B8860B"

headers_events = ["Date", "Event", "Host / Location", "Time", "Category", "Notes"]
ws_events.append(headers_events)
for i, w in enumerate([14, 40, 30, 22, 14, 40], 1):
    ws_events.column_dimensions[get_column_letter(i)].width = w

# Village events. Source: src/content/events/*.md. Sorted by date.
village_events = [
    ["2026-05-02", "Stories for Pups", "Kinderhook Memorial Library / 18 Hudson Street", "1–2 PM", "library", "Read aloud to a therapy dog."],
    ["2026-05-09", "Open Maker Hours — Mother's Day Card Making", "Super Stories", "10 AM – 12 PM", "community", ""],
    ["2026-05-09", "Travel Planning Made Easy", "Kinderhook Memorial Library / 18 Hudson Street", "2–3 PM", "library", ""],
    ["2026-05-16", "Festival of the Unknown", "Kinderhook Memorial Library / 18 Hudson Street", "11 AM – 2 PM", "library", "Cryptozoology, ufology, parapsychology lectures + documentary."],
    ["2026-05-23", "Village-Wide Spring Yard Sale", "Village of Kinderhook", "9 AM until sold out", "community", ""],
    ["2026-05-30", "Kinderhook Makers Market — Extended Market Day", "Kinderhook Farmers Market & Makers Market / Village Green", "8:30 AM – 2:00 PM", "community", "Extended market hours"],
    ["2026-05-30", "Fyfe & Drumms Muster & Parade", "Village of Kinderhook / Broad Street", "Noon", "community", ""],
    ["2026-05-30", "Modus Operandi — Opening Reception", "The School / Jack Shainman Gallery / 25 Broad Street", "2 PM", "art", "Group show: El Anatsui, Nick Cave, Faith Ringgold, et al."],
    ["2026-06-06", "Stories for Pups", "Kinderhook Memorial Library / 18 Hudson Street", "1–2 PM", "library", "Read aloud to a therapy dog."],
    ["2026-06-06", "Seen Scenes — Opening Reception", "Kinderhook Knitting Mill / Create Council on the Arts", "3 – 5 PM", "art", "Members Show 2026, on view June 5–28."],
    ["2026-06-06", "OK 5K", "Village of Kinderhook", "TBD", "community", ""],
    ["2026-06-06", "Persons of Color Cemetery Tour", "The Cultural Landscape Foundation", "TBD", "community", ""],
    ["2026-06-13", "Volunteer Fair", "Kinderhook Memorial Library / 18 Hudson Street", "10 AM – 2 PM", "library", "Library's inaugural volunteer fair — meet local nonprofits."],
    ["2026-06-20", "Rising Star Dance Academy Performance", "Village Green / At the Market", "11 AM", "community", "Live performance at the market."],
    ["2026-06-27", "Kinderhook Pride Parade", "Hudson Street to Kinderhook Village Square", "2 PM", "community", ""],
    ["2026-07-04", "People's Parade", "KBPA / Rothermel Park", "11:30 AM", "community", "Market extended to 1:30 PM"],
    ["2026-09-19", "Village-Wide Fall Yard Sale", "Village of Kinderhook", "9 AM until sold out", "community", ""],
    ["2026-10-10", "Fall Festival & Kinderhook Makers Market", "Kinderhook Farmers Market & Makers Market / Village Green", "8:30 AM – 2:00 PM", "community", "Extended market hours"],
    ["2026-10-31", "Final Market Day of 2026 Season", "Kinderhook Farmers Market / Village Green", "8:30 AM – 12:30 PM", "food", ""],
]
for e in village_events:
    ws_events.append(e)

style_headers(ws_events, len(headers_events))
style_data_rows(ws_events, len(headers_events))


# ─── Tab 6: Sponsors ───

ws_sponsors = wb.create_sheet("Sponsors")
ws_sponsors.sheet_properties.tabColor = "F26A2A"

headers_sponsors = ["Sponsor", "Website", "Logo File", "Tagline", "Notes"]
ws_sponsors.append(headers_sponsors)
for i, w in enumerate([40, 40, 36, 36, 40], 1):
    ws_sponsors.column_dimensions[get_column_letter(i)].width = w

# Sponsors. Source: src/data/sponsors.json.
sponsors_list = [
    ["Berkshire Hathaway HomeServices Blake Realtors", "https://www.bhhsblakerealtors.com/columbia",
     "/icons/sponsors/berkshire-hathaway.png", "Real estate · Columbia County",
     "Kinderhook office in the village since 1922."],
    ["Columbia County Tourism", "https://columbiacountytourism.org/",
     "/icons/sponsors/columbia-county-tourism.png", "Get out and explore", ""],
    ["Herrington's", "https://herringtons.com/",
     "/icons/sponsors/herringtons.png", "Lumber, hardware, and home goods",
     "Family-owned. Five locations across the Hudson Valley."],
    ["Julia Jayne Pilates", "https://www.juliajaynepilates.com/",
     "/icons/sponsors/julia-jayne-pilates.png", "Classical pilates · Broad Street",
     "Renovated barn studio on Broad Street."],
    ["Todd Farrell's Car Care Center", "https://www.toddfarrells.com/",
     "/icons/sponsors/todd-farrells.png", "Auto repair and tires · Hudson",
     "Family-run since 1995. Best of Columbia County winners."],
    ["Valkin Properties", "https://www.valkinproperties.com/",
     "/icons/sponsors/valkin-properties.png", "Real estate · Kinderhook & Valatie",
     "Bill Laraway's brokerage — rooted in Kinderhook and Valatie."],
]
for s in sponsors_list:
    ws_sponsors.append(s)

style_headers(ws_sponsors, len(headers_sponsors))
style_data_rows(ws_sponsors, len(headers_sponsors))


# ─── Tab 7: Site Config ───

ws4 = wb.create_sheet("Site Config")
ws4.sheet_properties.tabColor = "1A1A1A"

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 50
ws4.column_dimensions["C"].width = 40

# Header
ws4.append(["Setting", "Value", "Notes"])
style_headers(ws4, 3)

config_rows = [
    ["Season Start", "2026-05-02", "First market Saturday"],
    ["Season End", "2026-10-31", "Last market Saturday"],
    ["Market Day", "Saturday", ""],
    ["Start Time", "8:30 AM", ""],
    ["End Time", "12:30 PM", ""],
    ["Location Name", "Village Green", ""],
    ["Address", "Village Green, Broad Street & Albany Avenue, Kinderhook, NY 12106", ""],
    ["Coordinates", "42.3951, -73.6981", ""],
    ["Email", "khookfarmersmarket@icloud.com", ""],
    ["Facebook", "https://www.facebook.com/KinderhookFarmersMarket", ""],
    ["Instagram", "https://www.instagram.com/kinderhookfarmersmarket", ""],
    ["Sponsor", "Kinderhook Business and Professional Association", ""],
    ["Tagline", "Fresh local produce, artisan goods, and live music in the heart of the Hudson Valley.", ""],
]
for r in config_rows:
    ws4.append(r)

style_data_rows(ws4, 3)

# Bold the setting names
for row in range(2, 2 + len(config_rows)):
    ws4.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=11)
    ws4.cell(row=row, column=3).font = Font(name="Arial", italic=True, size=10, color="888888")


# ─── Tab 5: Instructions ───

ws5 = wb.create_sheet("How to Use")
ws5.sheet_properties.tabColor = "2AAFCE"
ws5.column_dimensions["A"].width = 80

instructions = [
    "KINDERHOOK FARMERS' MARKET - CONTENT MANAGEMENT SHEET",
    "",
    "HOW TO USE THIS SPREADSHEET",
    "=" * 50,
    "",
    "WEEKLY SCHEDULE TAB:",
    "  - Each row is one Saturday market day (pre-filled for the full season)",
    "  - Fill in the Theme, Music, Events, and Vendors for each week",
    "  - Set Status to 'Ready' when the week's info is complete",
    "  - After the website is updated, status changes to 'Published'",
    "  - Guest Vendors, Community Partners, and Regular Vendors are comma-separated lists",
    "  - Leave fields blank if not applicable (e.g., no music that week)",
    "",
    "VENDORS TAB:",
    "  - One row per vendor (both weekly regulars and rotating guests)",
    "  - Categories are comma-separated (e.g., 'produce, fruit, cider')",
    "  - Set Active to 'No' if a vendor is taking a break",
    "  - Set Featured to 'Yes' for vendors to highlight on the homepage",
    "",
    "RECIPES TAB:",
    "  - One row per recipe",
    "  - Ingredients and Instructions can have line breaks (Alt+Enter in Google Sheets)",
    "  - Produce Used links recipes to the seasonal guide",
    "  - Set Status to 'Ready' when a recipe is reviewed and good to publish",
    "",
    "SITE CONFIG TAB:",
    "  - Rarely needs updating - only for season dates, hours, or contact changes",
    "",
    "WORKFLOW:",
    "  1. Fill in the upcoming week's row on the Weekly Schedule tab",
    "  2. Set the Status to 'Ready'",
    "  3. Share with David or notify the web team",
    "  4. Website gets updated and status changes to 'Published'",
]

for i, line in enumerate(instructions, 1):
    cell = ws5.cell(row=i, column=1, value=line)
    if i == 1:
        cell.font = Font(name="Arial", bold=True, size=14, color="F26A2A")
    elif line.startswith("="):
        cell.font = Font(name="Arial", size=10, color="DDDDDD")
    elif line.endswith(":") and line == line.upper():
        cell.font = Font(name="Arial", bold=True, size=12, color="1A1A1A")
    elif line.startswith("  -"):
        cell.font = Font(name="Arial", size=10)
    else:
        cell.font = Font(name="Arial", size=11)


# Save
output_path = "/Users/davidnyman/Code/market/kinderhook-market-content.xlsx"
wb.save(output_path)
print(f"Spreadsheet saved to: {output_path}")
